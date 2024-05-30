from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User, Group
from .models import AdminProduct as Product, PlumberRequest, MainCategory, ProductEnquiry as Enquiry
import datetime, base64
import threading
import re

def remove_spcl_ch(code_string):return re.sub(r'\W+', '-', "-".join(map(str, code_string)).lower())

def get_role_name(request):
    role = request.user.groups.all()[0].name
    if role=="per_user":role = "User"
    elif role=="prime":role = "Prime"
    elif role=="plumber":role = "Plumber"
    elif role=="user_admin":role = "User Admin"
    return role

def session_products(request):
    prod_dict = []
    try:
        user = request.user.groups.all()[0].name
        if user == 'per_user':url = '/user/product-detail/'
        elif user == 'prime':url = '/prime/product-detail/'
        elif user == 'plumber':url = '/plumber/product-detail/'
        elif user == 'user_admin':url = '/user-admin/product-detail/'
    except: url = '/signin/?='
    
    
    for i in Product.objects.all():
        prod_dict.append({
            'pro_name':i.pro_name,
            'pro_image':i.pro_image,
            'url':url+i.pro_code
        })
    
    request.session["prod_dict"] = prod_dict
    
    return JsonResponse({"prod_dict":prod_dict})

def enquiry(request):
    user_name = request.POST.get('user_name')
    user_contact = request.POST.get('user_contact')
    product_name = request.POST.get('product_name')
    product_code = request.POST.get('product_code')
    user_city = request.POST.get('user_city')
    user_state = request.POST.get('user_state')
    user_message = request.POST.get('user_message')
    enquiry = {
        'request_id': remove_spcl_ch([datetime.datetime.now(),user_name,user_contact,user_city,user_state]),
        'user_name': user_name,
        'user_contact': user_contact,
        'product_name':product_name,
        'product_code':product_code,
        'user_city': user_city,
        'user_state': user_state,
        'user_message': user_message,
    }
    Enquiry.objects.create(**enquiry)
    return JsonResponse({'status':200})

def index(request):
    threading.Thread(target=session_products, args=(request,)).start()
    
    if request.user.is_authenticated:
        return redirect("auth_user")
    else:
        if request.method == 'POST':
            name = request.POST['name']
            contact = request.POST['contact']
            address = request.POST['address']        
            try:
                if request.POST['plumber_visit']=='on':plumber_visit = True
            except:plumber_visit = False
            if 'product_image' in request.FILES:
                image = base64.b64encode(request.FILES['product_image'].read()).decode('utf-8')
            else:image = ""
            id=str(str(datetime.datetime.now())+"-"+name+"-"+contact).strip()
            for i in id:
                if i.isalnum()==False and i!='-':
                    id = id.replace(i,'-')
            PlumberRequest.objects.create(
                request_id=id,
                name=name,
                contact=contact,
                address=address,
                plumber_visit=plumber_visit,
                product_image=image
            )
          
        trendings_products_one = Product.objects.filter(trending=True)[:4]
        trendings_products_two = Product.objects.filter(trending=True)[4:]
        
        gallery_products = Product.objects.filter(gallery_view=True)
        
        product_one = Product.objects.filter(trending=True).first()
        if product_one!=None:
            features = product_one.pro_features.split('\n')
            for j in range(len(features)):
                features[j] = features[j].strip()
            product_one.pro_features = features
        
        product_two = Product.objects.filter(trending=True).last()
        if product_two!=None:        
            features = product_two.pro_features.split('\n')
            for j in range(len(features)):
                features[j] = features[j].strip()
            product_two.pro_features = features
        
        return render(
            request,
            'home.html',
            {
                'main_categories':MainCategory.objects.all(),
                'trending_products_one':trendings_products_one,
                'gallery_products':Product.objects.filter(gallery_view=True)
            }
        )

def search(request):
    search = str(request.POST.get('search')).lower()
    try:
        user = request.user.groups.all()[0].name
        if user == 'per_user':url = '/user/product-detail/'
        elif user == 'prime':url = '/prime/product-detail/'
        elif user == 'plumber':url = '/plumber/product-detail/'
        elif user == 'user_admin':url = '/user-admin/product-detail/'
    except: url = '/signin'
    products = []
    if search!="":
        for i in Product.objects.all():
            if search in str(i.pro_name).lower():
                products.append({'name':i.pro_name,'image':i.pro_image,'url':url+i.product_id})
        return JsonResponse({'search':search,'status':200,'products':products})
    return JsonResponse({'status':400})

def auth_user(request):
    if request.user.is_authenticated:
        if request.user.groups.filter(name='super-admin').exists():return redirect('/admin')
        elif request.user.groups.filter(name='user_admin').exists():return redirect('user-admin-home')
        elif request.user.groups.filter(name='per_user').exists():return redirect('per_user')
        elif request.user.groups.filter(name='plumber').exists():return redirect('plumber')
        elif request.user.groups.filter(name='prime').exists():return redirect('prime')
        else:return redirect('signin')
    else:return redirect('signin')

def signup(request):
    if request.user.is_authenticated:return redirect("auth_user")
    elif request.method=="POST":
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        role = request.POST['role']
        if User.objects.filter(email=email).exists():messages.info(request, 'Email already exists')
        else:
            if role=='per_user':
                user = User.objects.create_user(username, email, password, is_active=True)
                user = authenticate(request, username=username, password=password)
                group = Group.objects.all().filter(name=role).first()
                user.groups.add(group)
                login(request, user)
            else:
                user = User.objects.create_user(username, email, password)
                user = authenticate(request, username=username, password=password)
                group = Group.objects.all().filter(name=role).first()
                user.groups.add(group)
                user = User.objects.get(username=username)
                user.is_active = False
                user.save()
                return redirect("signin")
            return redirect("auth_user")
        return redirect("signup")
    else:return render(request,'account/signup.html')

def signin(request):
    if request.user.is_authenticated:return redirect("auth_user")
    elif request.method == 'POST':
        username = str(request.POST["username"]).lower()
        password = request.POST["password"]
        if '@' in username:username = User.objects.get(email=username).username
        if User.objects.filter(username=username).exists():
            user = authenticate(request, username=username, password=password)
            if user and User.objects.get(username=username).is_active:
                login(request, user)
                return redirect("auth_user")
            elif User.objects.get(username=username).is_active==False:
                return render(request,'account/signin.html',{'message':'User need to be Approved'})
            else:
                return render(request,'account/signin.html',{'message':'Invalid User Credential'})
        else:return render(request,'account/signin.html',{'message':'User Not Found'})
    else:return render(request,'account/signin.html')

def signout(request):
    if request.user.is_authenticated:logout(request)
    return redirect('signin')

def error(request):
    import json
    from django.core.serializers import serialize, deserialize
    from .models import AdminProduct, SubCategory, MainCategory
    import csv

    product_file = "product.json"
    maincategory_file = "maincategory.json"
    subcategory_file = "subcategory.json"

    def handle():
        filename = 'mainc.csv'
        data = MainCategory.objects.all().values()
        headers = data[0].keys()
        with open(filename, 'w', newline='') as csvfile:
            csvwriter = csv.writer(csvfile)
            csvwriter.writerow(headers)
            for row in data:
                csvwriter.writerow(row.values())
        print(f'Data successfully exported to {filename}')

    def export_to_json():
        # product = AdminProduct.objects.all()
        # subcategory = SubCategory.objects.all()
        # maincategory = MainCategory.objects.all()

        # serialized_data_product = serialize('json', product)
        # serialized_data_subcategory = serialize('json', subcategory)
        # serialized_data_maincategory = serialize('json', maincategory)
        
        # with open(product_file, 'w') as json_file:
        #     json_file.write(serialized_data_product)

        # with open(subcategory_file, 'w') as json_file:
        #     json_file.write(serialized_data_subcategory)

        # with open(maincategory_file, 'w') as json_file:
        #     json_file.write(serialized_data_maincategory)
        pass

    def import_from_json():
        
        product_file = "product.json"
        subcategory_file = "subcategory.json"
        maincategory_file = "maincategory.json"
        
        with open(product_file, 'r') as json_file:
            json_data = json.load(json_file)
            
        for obj in deserialize('json', json_data):
            obj.save()
            
        with open(subcategory_file, 'r') as json_file:
            json_data = json.load(json_file)
            
        for obj in deserialize('json', json_data):
            obj.save()
        
        with open(maincategory_file, 'r') as json_file:
            json_data = json.load(json_file)
            
        for obj in deserialize('json', json_data):
            obj.save()

    def json_to_db_product():
        with open(product_file, 'r') as file:
            data = json.load(file)

        for i in data:
            obj = i['fields']
            print(obj['pro_name'])
            dic = {
                'pro_name' : obj['pro_name'],
                'pro_description' : obj['pro_description'],
                'pro_code' : obj['pro_code'],
                'pro_features' : obj['pro_features'],
                'pro_price' : obj['pro_price'],
                'pro_price_per_user' : obj['pro_price_per_user'],
                'pro_price_plumber' : obj['pro_price_plumber'],
                'pro_price_prime' : obj['pro_price_prime'],
                'main_category' : obj['main_category'],
                'sub_category' : obj['sub_category'],
                'trending' : obj['trending'],
                'gallery_view' : obj['gallery_view'],
                'pro_series' : obj['pro_series'],
                'product_id' : i['pk'],
                'pro_image' : obj['pro_image'],
            }
            AdminProduct.objects.create(**dic)
            
    def json_to_db_mainc():
        with open(maincategory_file, 'r') as file:
            data = json.load(file)

        for i in data:
            obj = i['fields']
            print(i['pk'])

            dic = {
                'cat_id' : re.sub(r'-+', '-', remove_spcl_ch(str(i['pk']).strip()+str(datetime.datetime.now()))),
                'name' : i['pk'],
                'description' : obj['description'],
                'image' : obj['image'],
            }
            MainCategory.objects.create(**dic)

    def json_to_db_subc():
        with open(subcategory_file, 'r') as file:
            data = json.load(file)

        for i in data:
            obj = i['fields']
            print(i['pk'])

            dic = {
                'cat_id' : re.sub(r'-+', '-', remove_spcl_ch(str(i['pk']).strip()+str(datetime.datetime.now()))),
                'name' : i['pk'],
                'main_category':obj['main_category'],
                'description' : obj['description'],
                'image' : obj['image'],
            }
            SubCategory.objects.create(**dic)

    def clear_db():
        Product.objects.all().delete()
        MainCategory.objects.all().delete()
        SubCategory.objects.all().delete()

    def excel_to_db():
        from openpyxl import load_workbook

        # Load the workbook
        file = 'product.xlsx'
        workbook = load_workbook(filename=file)

        # Select the active sheet
        sheet = workbook.active

        # Access data in the sheet
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # Display the data
        for row in data:
            row=list(row)
            pro_name = row[0]
            pro_description = row[1]
            pro_code = row[2]
            pro_features = row[3]
            pro_price = row[4]
            pro_price_per_user = row[5]
            pro_price_plumber = row[6]
            pro_price_prime = row[7]
            main_category = row[8]
            sub_category = row[9]
            trending = row[10]
            gallery_view = row[11]
            pro_series = [12]
            product_id = [13]
            pro_image = row[14]
            dic = {
                'pro_name' : row[0],
                'pro_description' : row[1],
                'pro_code' : row[2],
                'pro_features' : row[3],
                'pro_price' : row[4],
                'pro_price_per_user' : row[5],
                'pro_price_plumber' : row[6],
                'pro_price_prime' : row[7],
                'main_category' : row[8],
                'sub_category' : row[9],
                'trending' : row[10],
                'gallery_view' : row[11],
                'pro_series' : row[12],
                'product_id' : row[13],
                'pro_image' : row[14],
            }
            print(dic)
            if row[0]=='pro_name':continue
            else:
                Product.objects.create(**dic)

    def repair():
        for i in Product.objects.all():
            if i.trending==0:
                i.trending = False
            elif i.trending==1:
                i.trending = True
            if i.gallery_view==0:
                i.gallery_view = False
            elif i.gallery_view==1:
                i.gallery_view = True
            i.save()
            print("Name: ",i.pro_name,"Trending: ",i.trending,"Gallery :",i.gallery_view)
    
    repair()
    # excel_to_db()
    # json_to_db_product()
    # json_to_db_mainc()
    # json_to_db_subc()
    # clear_db()
    # export_to_json()
    # import_from_json()
    # handle()

    return render(request, 'error.html')

def delete_account(request):    
    if request.user.is_authenticated and not request.user.is_superuser:
        user = User.objects.get(email=request.user.email)
        user.delete()
        return redirect('signin')
    else:return redirect('signin')

